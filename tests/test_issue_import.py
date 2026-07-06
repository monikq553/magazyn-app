import io
import unittest
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import MagicMock, patch

from openpyxl import Workbook, load_workbook

import app as warehouse_app
from general_import import parse_workbook
from issue_import import (
    issue_history_pdf,
    issue_history_xlsx,
    issue_mapping,
    issue_sheet_selected,
    normalize_issue_row,
    validate_issue_row,
)


def workbook_bytes(rows, sheet_name="Wydania"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def valid_data(package_number="PACZKA-1"):
    return {
        "date": "2026-07-06",
        "doc_number": "WZ/100",
        "contractor": "Klient Testowy",
        "product_name": "Próbka",
        "dimension": "20x40",
        "species": "Dąb",
        "package_number": package_number,
        "package_id": "",
        "qty": 2.0,
        "unit": "szt",
        "warehouse": "Inne",
        "notes": "Test",
    }


class IssueImportTests(unittest.TestCase):
    def test_issue_sheet_and_all_requested_columns_are_recognized(self):
        parsed = parse_workbook(
            workbook_bytes(
                [
                    [
                        "Data",
                        "Dokument",
                        "Odbiorca",
                        "Produkt",
                        "Wymiar",
                        "Gatunek",
                        "Numer paczki",
                        "Ilość",
                        "Jednostka",
                        "Magazyn",
                        "Uwagi",
                    ],
                    [
                        "2026-07-06",
                        "WZ/100",
                        "Klient",
                        "Próbka",
                        "20x40",
                        "Dąb",
                        "PACZKA-1",
                        2,
                        "szt",
                        "Inne",
                        "Test",
                    ],
                ]
            )
        )
        sheet = parsed[0]
        mapping = issue_mapping(sheet["columns"])
        data = normalize_issue_row(sheet["rows"][0]["source_data"], mapping)

        self.assertTrue(issue_sheet_selected(sheet))
        self.assertEqual(data["doc_number"], "WZ/100")
        self.assertEqual(data["contractor"], "Klient")
        self.assertEqual(data["dimension"], "20x40")
        self.assertEqual(data["species"], "Dąb")
        self.assertEqual(data["notes"], "Test")

    def test_invalid_issue_row_reports_quantity_unit_warehouse_and_date(self):
        data = valid_data()
        data.update(
            {
                "date": "31-31-2026",
                "qty": "-2",
                "unit": "metr",
                "warehouse": "Nieznany",
            }
        )
        errors = validate_issue_row(
            data,
            warehouse_app.UNITS,
            warehouse_app.WAREHOUSES,
        )

        self.assertTrue(any("Data" in error for error in errors))
        self.assertTrue(any("Ilość" in error for error in errors))
        self.assertIn("Nieprawidłowa jednostka.", errors)
        self.assertIn("Nieprawidłowy magazyn.", errors)

    def test_issue_context_resolves_existing_package_and_available_stock(self):
        cursor = MagicMock()
        cursor.fetchone.side_effect = [
            (3, 20.0, "szt", 4.0, 23.0),
            (8, "PACZKA-1", 10.0),
        ]
        data = valid_data()

        errors, context = warehouse_app.issue_import_row_context(
            cursor,
            data,
            allow_general_stock=False,
        )

        self.assertEqual(errors, [])
        self.assertEqual(context["product"][0], 3)
        self.assertEqual(context["package"][0], 8)
        self.assertEqual(data["package_id"], 8)

    def test_issue_without_package_obeys_admin_setting(self):
        for allowed, expected_error in ((False, True), (True, False)):
            with self.subTest(allowed=allowed):
                cursor = MagicMock()
                cursor.fetchone.return_value = (3, 20.0, "szt", 4.0, 23.0)
                data = valid_data(package_number="")
                errors, context = warehouse_app.issue_import_row_context(
                    cursor,
                    data,
                    allow_general_stock=allowed,
                )
                self.assertEqual(bool(errors), expected_error)
                self.assertIsNone(context["package"])

    def test_one_and_multiple_rows_create_one_document_and_multiple_items(self):
        cursor = MagicMock()
        cursor.rowcount = 1
        cursor.fetchone.side_effect = [(100,), (200,), (201,)]
        document_cache = {}
        context = {
            "product": (3, 20.0, "szt", 4.0, 23.0),
            "package": (8, "PACZKA-1", 10.0),
            "existing_item": None,
        }
        row_one = (1, "Wydania", 2, {}, valid_data(), None, "new", True, [])
        row_two = (2, "Wydania", 3, {}, valid_data(), None, "new", True, [])

        first = warehouse_app.apply_issue_import_row(
            cursor, 9, row_one, valid_data(), context, document_cache
        )
        second = warehouse_app.apply_issue_import_row(
            cursor, 9, row_two, valid_data(), context, document_cache
        )

        self.assertEqual((first, second), ("added", "added"))
        document_inserts = [
            call
            for call in cursor.execute.call_args_list
            if "INSERT INTO issue_docs" in call.args[0]
        ]
        item_inserts = [
            call
            for call in cursor.execute.call_args_list
            if "INSERT INTO issue_items" in call.args[0]
        ]
        self.assertEqual(len(document_inserts), 1)
        self.assertEqual(len(item_inserts), 2)

    def test_duplicate_update_changes_existing_item_instead_of_adding_second(self):
        cursor = MagicMock()
        cursor.rowcount = 1
        cursor.fetchone.return_value = (
            "2026-07-01",
            "Stary klient",
            "Inne",
            "WZ/100",
        )
        data = valid_data()
        row = (
            1,
            "Wydania",
            2,
            {},
            data,
            {"id": 50, "label": "WZ/100"},
            "update",
            True,
            [],
        )
        context = {
            "product": (3, 20.0, "szt", 4.0, 23.0),
            "package": (8, "PACZKA-1", 10.0),
            "existing_item": (70, 1.0, 8, "PACZKA-1", "", "", ""),
        }

        result = warehouse_app.apply_issue_import_row(
            cursor, 9, row, data, context, {}
        )

        self.assertEqual(result, "updated")
        self.assertTrue(
            any(
                "UPDATE issue_items" in call.args[0]
                for call in cursor.execute.call_args_list
            )
        )
        self.assertFalse(
            any(
                "INSERT INTO issue_items" in call.args[0]
                for call in cursor.execute.call_args_list
            )
        )

    def test_undo_restores_product_package_and_marks_created_document_void(self):
        cursor = MagicMock()
        cursor.rowcount = 1
        cursor.fetchall.return_value = [
            (
                1,
                10,
                "added_item",
                100,
                200,
                3,
                8,
                2.0,
                {"doc_created": True},
            )
        ]
        cursor.fetchone.side_effect = [(None,), (2.0,)]

        warehouse_app.undo_issue_import(cursor, 9, "admin@example.com")

        sql = "\n".join(call.args[0] for call in cursor.execute.call_args_list)
        self.assertIn("UPDATE products SET qty=qty+%s", sql)
        self.assertIn("UPDATE packages", sql)
        self.assertIn("voided_at=NOW()", sql)
        self.assertIn("status='undone'", sql)

    def test_undo_refuses_document_already_cancelled_elsewhere(self):
        cursor = MagicMock()
        cursor.fetchall.return_value = [
            (
                1,
                10,
                "added_item",
                100,
                200,
                3,
                8,
                2.0,
                {"doc_created": True},
            )
        ]
        cursor.fetchone.return_value = (datetime(2026, 7, 6, 13, 0),)

        with self.assertRaisesRegex(ValueError, "już anulowany"):
            warehouse_app.undo_issue_import(cursor, 9, "admin@example.com")

    def test_history_exports_are_valid_xlsx_and_pdf(self):
        rows = [
            (
                9,
                "wydania.xlsx",
                "admin@example.com",
                "completed",
                {"issues": 3, "updated": 1, "skipped": 2, "errors": 0},
                [],
                datetime(2026, 7, 6, 12, 30, tzinfo=timezone.utc),
            )
        ]
        xlsx = issue_history_xlsx(rows)
        pdf = issue_history_pdf(rows)

        workbook = load_workbook(io.BytesIO(xlsx))
        sheet = workbook["Historia importów"]
        self.assertEqual(sheet["B2"].value, "wydania.xlsx")
        self.assertEqual(sheet["E2"].value, 3)
        self.assertTrue(pdf.startswith(b"%PDF"))
        self.assertGreater(len(pdf), 2000)

        unsafe_rows = [
            (
                10,
                "=HYPERLINK(\"https://example.test\")",
                "@user",
                "draft",
                {},
                [],
                datetime(2026, 7, 6, tzinfo=timezone.utc),
            )
        ]
        safe_book = load_workbook(io.BytesIO(issue_history_xlsx(unsafe_rows)))
        self.assertTrue(safe_book["Historia importów"]["B2"].value.startswith("'="))
        self.assertTrue(safe_book["Historia importów"]["C2"].value.startswith("'@"))

    def test_preview_keeps_mapping_editing_add_delete_and_confirm_separate(self):
        run = (
            9,
            "wydania.xlsx",
            "admin@example.com",
            "draft",
            [
                {
                    "name": "Wydania",
                    "columns": ["Data", "Dokument", "Produkt"],
                    "mapping": {
                        "date": "Data",
                        "doc_number": "Dokument",
                        "product_name": "Produkt",
                    },
                    "row_count": 1,
                    "selected": True,
                }
            ],
            {},
            [],
            datetime(2026, 7, 6, 12, 0),
            None,
            None,
            None,
        )
        rows = [
            (
                1,
                "Wydania",
                2,
                {"Data": "2026-07-06"},
                valid_data(),
                None,
                "new",
                True,
                [],
            )
        ]
        cursor = MagicMock()
        cursor.fetchone.side_effect = [run, ("false",)]
        cursor.fetchall.side_effect = [
            rows,
            [(8, "PACZKA-1", 10.0, "Inne", "Próbka", "szt")],
        ]
        connection = MagicMock()
        connection.cursor.return_value = cursor

        with warehouse_app.app.test_request_context("/import-wydan/9"):
            warehouse_app.session["user"] = "admin@example.com"
            warehouse_app.session["role"] = "admin"
            with patch.object(warehouse_app, "db", return_value=connection):
                html = warehouse_app.issue_import_preview(9)

        self.assertIn('/import-wydan/9/mapping', html)
        self.assertIn('/import-wydan/9/prepare', html)
        self.assertIn('/import-wydan/9/rows', html)
        self.assertIn('/import-wydan/9/rows/1/delete', html)
        self.assertIn('/import-wydan/9/confirm', html)
        self.assertIn('row_1_package_id', html)

    def test_migration_contains_issue_import_history_effects_and_item_fields(self):
        source = Path(warehouse_app.__file__).read_text(encoding="utf-8")
        for fragment in (
            "CREATE TABLE IF NOT EXISTS issue_imports",
            "CREATE TABLE IF NOT EXISTS issue_import_rows",
            "CREATE TABLE IF NOT EXISTS issue_import_effects",
            "ADD COLUMN IF NOT EXISTS dimension",
            "ADD COLUMN IF NOT EXISTS species",
            "ADD COLUMN IF NOT EXISTS notes",
            "REFERENCES issue_items(id) ON DELETE SET NULL",
        ):
            self.assertIn(fragment, source)


if __name__ == "__main__":
    unittest.main()
