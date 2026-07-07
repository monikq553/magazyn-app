import io
import os
from datetime import datetime
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Image as ReportLabImage,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from general_import import (
    FIELD_LABELS,
    guess_mapping,
    normalize_row,
    normalized_text,
    parse_date,
    parse_number,
)


ISSUE_IMPORT_FIELDS = (
    "date",
    "doc_number",
    "contractor",
    "product_name",
    "dimension",
    "species",
    "package_number",
    "qty",
    "unit",
    "warehouse",
    "notes",
)

ISSUE_IMPORT_LABELS = {
    field: FIELD_LABELS.get(field, field)
    for field in ISSUE_IMPORT_FIELDS
}


def excel_safe_text(value):
    text = str(value or "")
    return "'" + text if text.startswith(("=", "+", "-", "@")) else text


def issue_sheet_selected(sheet):
    if sheet.get("entity_type") in {"issue", "document"}:
        return True
    name = normalized_text(sheet.get("name"))
    return any(token in name for token in ("wydan", "wz", "issued"))


def issue_mapping(columns):
    return {
        field: column
        for field, column in guess_mapping(columns, "issue").items()
        if field in ISSUE_IMPORT_FIELDS
    }


def normalize_issue_row(source_data, mapping):
    data = normalize_row("issue", source_data, mapping)
    return {field: data.get(field, "") for field in ISSUE_IMPORT_FIELDS}


def validate_issue_row(data, units, warehouses):
    errors = []
    for field in (
        "date",
        "doc_number",
        "contractor",
        "product_name",
        "qty",
        "unit",
        "warehouse",
    ):
        if not str(data.get(field) or "").strip():
            errors.append(f"{ISSUE_IMPORT_LABELS[field]}: pole wymagane.")
    try:
        if data.get("date"):
            data["date"] = parse_date(data["date"])
    except ValueError as exc:
        errors.append(str(exc))
    try:
        if data.get("qty") not in ("", None):
            data["qty"] = parse_number(data["qty"], "Ilość", allow_zero=False)
    except ValueError as exc:
        errors.append(str(exc))
    unit = str(data.get("unit") or "").strip()
    warehouse = str(data.get("warehouse") or "").strip()
    if unit and unit not in units:
        errors.append("Nieprawidłowa jednostka.")
    if warehouse and warehouse not in warehouses:
        errors.append("Nieprawidłowy magazyn.")
    if len(str(data.get("doc_number") or "")) > 100:
        errors.append("Numer dokumentu może mieć maksymalnie 100 znaków.")
    if len(str(data.get("package_number") or "")) > 100:
        errors.append("Numer paczki może mieć maksymalnie 100 znaków.")
    if len(str(data.get("notes") or "")) > 1000:
        errors.append("Uwagi mogą mieć maksymalnie 1000 znaków.")
    return errors


def history_rows_to_dicts(rows):
    result = []
    for row in rows:
        summary = row[4] or {}
        errors = row[5] or []
        created_at = row[6]
        if isinstance(created_at, datetime) and created_at.tzinfo is not None:
            created_at = created_at.replace(tzinfo=None)
        result.append(
            {
                "date": created_at,
                "filename": excel_safe_text(row[1]),
                "user": excel_safe_text(row[2]),
                "status": {
                    "completed": "zakończony",
                    "undone": "cofnięty",
                    "draft": "wersja robocza",
                }.get(row[3], excel_safe_text(row[3])),
                "issues": int(summary.get("issues", 0)),
                "errors": int(summary.get("errors", 0)) + len(errors),
                "skipped": int(summary.get("skipped", 0)),
                "updated": int(summary.get("updated", 0)),
            }
        )
    return result


def issue_history_xlsx(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Historia importów"
    headers = [
        "Data",
        "Plik",
        "Użytkownik",
        "Status",
        "Wydania",
        "Błędy",
        "Pominięte",
        "Zaktualizowane",
    ]
    sheet.append(headers)
    for item in history_rows_to_dicts(rows):
        sheet.append(
            [
                item["date"],
                item["filename"],
                item["user"],
                item["status"],
                item["issues"],
                item["errors"],
                item["skipped"],
                item["updated"],
            ]
        )
    header_fill = PatternFill("solid", fgColor="185C43")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    widths = (20, 34, 30, 16, 12, 12, 12, 16)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet["A"][1:]:
        if isinstance(cell.value, datetime):
            cell.number_format = "yyyy-mm-dd hh:mm"
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _pdf_fonts():
    regular_candidates = (
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        r"C:\Windows\Fonts\arial.ttf",
    )
    bold_candidates = (
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        r"C:\Windows\Fonts\arialbd.ttf",
    )
    regular = next((path for path in regular_candidates if os.path.isfile(path)), None)
    bold = next((path for path in bold_candidates if os.path.isfile(path)), None)
    if regular and bold:
        if "IssueImportSans" not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(TTFont("IssueImportSans", regular))
            pdfmetrics.registerFont(TTFont("IssueImportSans-Bold", bold))
        return "IssueImportSans", "IssueImportSans-Bold"
    return "Helvetica", "Helvetica-Bold"


def _primadera_pdf_logo(width=58 * mm, align="CENTER"):
    path = os.path.join(os.path.dirname(__file__), "static", "primadera-logo.png")
    if not os.path.isfile(path):
        return None
    image = ReportLabImage(path, width=width, height=width * 86 / 392)
    image.hAlign = align
    return image


def issue_history_pdf(rows):
    body_font, bold_font = _pdf_fonts()
    output = io.BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="Historia importów wydań",
        author="Primadera",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "IssueImportTitle",
        parent=styles["Title"],
        fontName=bold_font,
        fontSize=17,
        leading=21,
        textColor=colors.HexColor("#185C43"),
        alignment=TA_CENTER,
    )
    cell_style = ParagraphStyle(
        "IssueImportCell",
        parent=styles["Normal"],
        fontName=body_font,
        fontSize=7.5,
        leading=9,
    )
    header_style = ParagraphStyle(
        "IssueImportHeader",
        parent=cell_style,
        fontName=bold_font,
        textColor=colors.white,
        alignment=TA_CENTER,
    )
    table_data = [
        [
            Paragraph(label, header_style)
            for label in (
                "Data",
                "Plik",
                "Użytkownik",
                "Status",
                "Wydania",
                "Błędy",
                "Pominięte",
                "Zaktualizowane",
            )
        ]
    ]
    for item in history_rows_to_dicts(rows):
        date_value = item["date"]
        if hasattr(date_value, "strftime"):
            date_value = date_value.strftime("%Y-%m-%d %H:%M")
        table_data.append(
            [
                Paragraph(escape(str(date_value or "")), cell_style),
                Paragraph(escape(str(item["filename"])), cell_style),
                Paragraph(escape(str(item["user"])), cell_style),
                Paragraph(escape(str(item["status"])), cell_style),
                Paragraph(str(item["issues"]), cell_style),
                Paragraph(str(item["errors"]), cell_style),
                Paragraph(str(item["skipped"]), cell_style),
                Paragraph(str(item["updated"]), cell_style),
            ]
        )
    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[31 * mm, 53 * mm, 48 * mm, 27 * mm, 20 * mm, 20 * mm, 22 * mm, 28 * mm],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#185C43")),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#C7D5CF")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story = []
    logo = _primadera_pdf_logo()
    if logo:
        story.extend([logo, Spacer(1, 4 * mm)])
    story.extend([
        Paragraph("Historia importów wydań towaru", title_style),
        Spacer(1, 5 * mm),
        table,
    ])
    document.build(story)
    return output.getvalue()
